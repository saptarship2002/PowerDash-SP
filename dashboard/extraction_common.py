"""Extract data/Common Indicators.xlsx (35 DISCOMs, ~12 states, 5 FYs each) into data/discoms2.json.

Unlike the old workbook (extraction3.py), every sheet shares one column layout, but each
sheet has a repeated block per financial year (its own 'Year' header row + data rows).
This walks each sheet generically, splitting on those header rows, and normalizes the
~20 raw indicator-name variants across states into 8 canonical indicators.
"""
import json, re
import openpyxl

wb = openpyxl.load_workbook('data/Common Indicators.xlsx', data_only=True)

STATE_MAP = {
    'MAHARASHTRA': 'Maharashtra', 'GUJARAT': 'Gujarat', 'RAJASTHAN': 'Rajasthan',
    'RAJ': 'Rajasthan', 'ODISHA': 'Odisha', 'TELANGANA': 'Telangana',
    'MADHYA PRADESH': 'Madhya Pradesh', 'KARNATAKA': 'Karnataka', 'TAMIL NADU': 'Tamil Nadu',
    'BIHAR': 'Bihar', 'WEST BENGAL': 'West Bengal', 'UP': 'Uttar Pradesh', 'AP': 'Andhra Pradesh',
}

STATE_ORDER = ['Maharashtra', 'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Odisha', 'Telangana',
               'Karnataka', 'Tamil Nadu', 'Bihar', 'West Bengal', 'Uttar Pradesh', 'Andhra Pradesh']

# Sheets with an obviously non-standard short code (everything else is derived from the
# part of the sheet name before the comma).
SHORT_OVERRIDES = {
    'JdVVNL,RAJ': 'JdVVNL',
}

# Uttar Pradesh's 5 sheets have a blank cell where every other sheet in the workbook carries the
# licensee's full legal name (row 1, col A) — confirmed against the source file, not a parsing
# gap. Filled in from UPPCL's own site (https://www.uppcl.org/uppcl/en/article/discoms) since the
# workbook itself has nothing to extract here.
FULL_NAME_OVERRIDES = {
    'PVVNL, UP': 'Pashchimanchal Vidyut Vitaran Nigam Limited',
    'MVVNL,UP': 'Madhyanchal Vidyut Vitaran Nigam Limited',
    'DVVNL,UP': 'Dakshinanchal Vidyut Vitaran Nigam Limited',
    'PuVVNL,UP': 'Purvanchal Vidyut Vitaran Nigam Limited',
    'KESCO,UP': 'Kanpur Electricity Supply Company Limited',
}

EXCLUDE_SHEETS = {'color scheme', 'verification'}

YEAR_RE = re.compile(r'^\d{4}-\d{2,4}$')

# ---- indicator normalization -------------------------------------------------------
CANONICAL = {
    'SAIDI': {'group': 'Reliability', 'unit': 'hours'},
    'SAIFI': {'group': 'Reliability', 'unit': 'count'},
    'MAIFI': {'group': 'Reliability', 'unit': 'count'},
    'CAIDI': {'group': 'Reliability', 'unit': 'hours'},
    'VOLTAGE VARIATION': {'group': 'Power Quality', 'unit': 'pct'},
    'HARMONICS': {'group': 'Power Quality', 'unit': 'pct'},
    'TRANSFORMER FAILURE': {'group': 'Service', 'unit': 'pct'},
    'BILLING COMPLAINT RESOLUTION': {'group': 'Consumer', 'unit': 'pct'},
}
CANONICAL_ORDER = list(CANONICAL.keys())


def normalize_indicator(raw):
    """('TRANSFORMER FAILURE', None|'Urban'|'Rural') from any raw indicator label."""
    if not raw:
        return None, None
    u = re.sub(r'\s+', ' ', raw.strip().upper())
    subtype = None
    if 'URBAN' in u:
        subtype = 'Urban'
    elif 'RURAL' in u:
        subtype = 'Rural'
    if 'SAIDI' in u:
        return 'SAIDI', subtype
    if 'SAIFI' in u:
        return 'SAIFI', subtype
    if 'MAIFI' in u:
        return 'MAIFI', subtype
    if 'CAIDI HT' in u:
        return None, None
    if 'CAIDI' in u:
        return 'CAIDI', subtype
    if 'VOLTAGE VARIATION' in u:
        return 'VOLTAGE VARIATION', subtype
    if 'HARMONICS' in u:
        return 'HARMONICS', subtype
    if 'TRANSFORMER FAILURE' in u or 'DT FAILURE' in u:
        return 'TRANSFORMER FAILURE', subtype
    if 'BILLING' in u:
        return 'BILLING COMPLAINT RESOLUTION', subtype
    return None, None


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def is_na(x):
    return x is None or str(x).strip().upper() in ('N/A', 'NA', '')


def to_bool_yn(x):
    if x is None:
        return None
    s = str(x).strip().upper()
    if s == 'YES':
        return True
    if s == 'NO':
        return False
    return None


def normalize_value(canon_key, raw_val, meaning):
    """Return (numeric_value, unit_note) for a canonical indicator's raw reported value."""
    if raw_val is None or isinstance(raw_val, str):
        return None, None
    val = float(raw_val)
    m = (meaning or '').lower()
    note = 'per quarter' if 'quarter' in m else None
    if canon_key in ('SAIDI', 'CAIDI'):
        if 'hour' in m or 'hr' in m:
            return round(val, 3), note
        return round(val / 60, 3), note  # default: reported in minutes
    if canon_key in ('SAIFI', 'MAIFI'):
        return round(val, 3), note
    # percentage indicators
    return round(val, 2), note


def parse_sheet(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))
    full_name = clean(rows[0][0]) or FULL_NAME_OVERRIDES.get(sheet_name)
    code, _, state_part = sheet_name.partition(',')
    code = code.strip()
    state = STATE_MAP.get(state_part.strip().upper(), state_part.strip() or 'Unknown')
    short_name = SHORT_OVERRIDES.get(sheet_name, code)

    blocks = {}  # year -> {'regulation': str, 'indicators': [...]}
    cur_reg = []
    i = 1
    n = len(rows)
    while i < n:
        r = rows[i]
        if r is None or all(c is None for c in r):
            i += 1
            continue
        c0 = clean(r[0])
        if c0 == 'Year':
            j = i + 1
            data = []
            while j < n:
                rj = rows[j]
                if rj is None or all(c is None for c in rj):
                    j += 1
                    continue
                c0j = clean(rj[0])
                if c0j == 'Year':
                    break
                if c0j and YEAR_RE.match(str(c0j)):
                    data.append(rj)
                    j += 1
                else:
                    break
            if data:
                year = clean(data[0][0])
                entries = []
                for rj in data:
                    raw_indicator = clean(rj[2])
                    canon, subtype = normalize_indicator(raw_indicator)
                    reported_raw = rj[7]
                    reported_meaning = clean(rj[8])
                    value = None
                    unit_note = None
                    if canon:
                        value, unit_note = normalize_value(canon, reported_raw, reported_meaning)
                    entries.append({
                        'raw_indicator': raw_indicator,
                        'canonical': canon,
                        'subtype': subtype,
                        'indicator_type': clean(rj[1]),
                        'indicator_meaning': clean(rj[3]),
                        'standard_specified': clean(rj[4]),
                        'benchmark': clean(rj[5]),
                        'benchmark_meaning': clean(rj[6]),
                        'reported_raw': clean(reported_raw),
                        'reported_meaning': reported_meaning,
                        'value': value,
                        'unit_note': unit_note,
                        'comparison_possible': to_bool_yn(rj[9]),
                        'standard_met': to_bool_yn(rj[10]),
                        'reason_not_comparable': clean(rj[11]) if len(rj) > 11 else None,
                    })
                blocks[year] = {'regulation': ' | '.join(cur_reg), 'indicators': entries}
                cur_reg = []
            i = j
        else:
            cur_reg.append(str(c0))
            i += 1

    return {
        'sheet': sheet_name,
        'full_name': full_name,
        'short_name': short_name,
        'state': state,
        'years': blocks,
    }


def collapse_canonical(entries):
    """Average Urban/Rural (and any duplicate) rows of the same canonical indicator
    into one value per canonical indicator, keeping the richest metadata available."""
    by_canon = {}
    for e in entries:
        if not e['canonical']:
            continue
        by_canon.setdefault(e['canonical'], []).append(e)
    out = {}
    for canon, es in by_canon.items():
        vals = [e['value'] for e in es if e['value'] is not None]
        rep = max(es, key=lambda e: 0 if e['value'] is None else 1)
        out[canon] = {
            'value': round(sum(vals) / len(vals), 3) if vals else None,
            'subtypes': {e['subtype'] or 'default': e['value'] for e in es},
            'indicator_meaning': rep['indicator_meaning'],
            'standard_specified': rep['standard_specified'],
            'benchmark': rep['benchmark'],
            'benchmark_meaning': rep['benchmark_meaning'],
            'reported_meaning': rep['reported_meaning'],
            'unit_note': rep['unit_note'],
            'comparison_possible': any(e['comparison_possible'] for e in es if e['comparison_possible'] is not None) if any(e['comparison_possible'] is not None for e in es) else None,
            'standard_met': all(e['standard_met'] for e in es if e['standard_met'] is not None) if any(e['standard_met'] is not None for e in es) else None,
            'reason_not_comparable': rep['reason_not_comparable'],
        }
    return out


def score_year(canon_map):
    n = len(CANONICAL_ORDER)
    present = [canon_map.get(k) for k in CANONICAL_ORDER]
    has_standard = sum(1 for c in present if c and not is_na(c['standard_specified']))
    reported = sum(1 for c in present if c and c['value'] is not None)
    claimed_comparable = [c for c in present if c and c['comparison_possible']]
    comparable_inds = [c for c in claimed_comparable if c['value'] is not None]
    comparable = len(comparable_inds)
    phantom = len(claimed_comparable) - comparable
    met = sum(1 for c in comparable_inds if c['standard_met'])
    composite = (
        (100 * has_standard / n) * 0.25 + (100 * reported / n) * 0.25 +
        (100 * comparable / n) * 0.30 +
        ((100 * met / len(comparable_inds)) if comparable_inds else 0) * 0.20
    )
    return {
        'total_indicators': n,
        'standards_available_pct': round(100 * has_standard / n, 1),
        'data_reported_pct': round(100 * reported / n, 1),
        'comparable_pct': round(100 * comparable / n, 1),
        'compliance_pct': round(100 * met / len(comparable_inds), 1) if comparable_inds else None,
        'composite_score': round(composite, 1),
        'grade': 'A' if composite >= 70 else ('B' if composite >= 45 else 'C'),
        'phantom_comparable_count': phantom,
        'indicators_reported': reported,
    }


discoms = []
for sheet_name in wb.sheetnames:
    if sheet_name in EXCLUDE_SHEETS:
        continue
    parsed = parse_sheet(sheet_name, wb[sheet_name])
    by_year = {}
    for year, block in parsed['years'].items():
        canon_map = collapse_canonical(block['indicators'])
        by_year[year] = {
            'regulation': block['regulation'],
            'indicators': canon_map,
            'scoring': score_year(canon_map),
        }
    discoms.append({
        'sheet': parsed['sheet'],
        'full_name': parsed['full_name'],
        'short_name': parsed['short_name'],
        'state': parsed['state'],
        'years': by_year,
    })

all_years = sorted({y for d in discoms for y in d['years']}, reverse=True)

with open('data/discoms2.json', 'w', encoding='utf-8') as f:
    json.dump({
        'state_order': STATE_ORDER,
        'canonical_indicators': CANONICAL,
        'canonical_order': CANONICAL_ORDER,
        'years': all_years,
        'discoms': discoms,
    }, f, indent=2, ensure_ascii=False)

print(f'Extracted {len(discoms)} DISCOMs across {len({d["state"] for d in discoms})} states, years={all_years}')
for d in discoms:
    y0 = all_years[0]
    sc = d['years'].get(y0, {}).get('scoring')
    print(f"  {d['short_name']:10} ({d['state']:16}) {y0}: score={sc['composite_score'] if sc else 'n/a'}")
