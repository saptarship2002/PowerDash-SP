"""Extract data/State specific Indicators.xlsx into ui/public/data/state_specific.json.

This is the Standards-of-Performance (SoP) counterpart to Common Indicators.xlsx: consumer
service indicators (fuse-off calls, line breakdowns, meter/billing complaints, new-connection
timelines, etc.) rather than SAIDI/SAIFI/power-quality. Unlike Common Indicators.xlsx, the raw
indicator names are too varied (dozens, state-specific) to normalize into a small canonical set,
so each DISCOM/year just keeps its raw indicator list exactly as reported — no derived score or
percentage of any kind. If a compliance score is wanted later, its formula needs sign-off first.

Two sheet shapes:
  - Per-DISCOM sheets (19 of them): same repeated-Year-block layout as Common Indicators.xlsx,
    but column count/order varies slightly sheet to sheet, so columns are located by header name
    rather than fixed index.
  - Per-state "framework only" sheets (UP, TAMIL NADU, KARNATAKA, ANDHRA PRADESH, WEST BENGAL):
    no Year column, no reported figures at all — just the regulation's named indicators,
    standards and benchmarks with nothing to compare them against.

Excludes the workbook's `verification` and `calculations` sheets — internal QA working notes,
not published indicator data.
"""
import json, re
import openpyxl

wb = openpyxl.load_workbook('data/State specific Indicators.xlsx', data_only=True)

STATE_MAP = {
    'MAHARASHTRA': 'Maharashtra', 'GUJARAT': 'Gujarat', 'RAJASTHAN': 'Rajasthan',
    'ODISHA': 'Odisha', 'TELANGANA': 'Telangana', 'MADHYA PRADESH': 'Madhya Pradesh',
    'KARNATAKA': 'Karnataka', 'TAMIL NADU': 'Tamil Nadu', 'BIHAR': 'Bihar',
    'WEST BENGAL': 'West Bengal', 'UP': 'Uttar Pradesh', 'UTTAR PRADESH': 'Uttar Pradesh',
    'AP': 'Andhra Pradesh', 'ANDHRA PRADESH': 'Andhra Pradesh',
}
STATE_ORDER = ['Maharashtra', 'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Odisha', 'Telangana',
               'Karnataka', 'Tamil Nadu', 'Bihar', 'West Bengal', 'Uttar Pradesh', 'Andhra Pradesh']

EXCLUDE_SHEETS = {'verification', 'calculations'}
FRAMEWORK_ONLY_SHEETS = {'UP', 'TAMIL NADU', 'KARNATAKA', 'ANDHRA PRADESH', 'WEST BENGAL'}

YEAR_RE = re.compile(r'^\d{4}-\d{2,4}$')


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_bool_yn(x):
    if x is None:
        return None
    s = str(x).strip().upper()
    if s == 'YES':
        return True
    if s == 'NO':
        return False
    return None


def to_number(x):
    if x is None or isinstance(x, str):
        return None
    return round(float(x), 2)


def header_cols(header_row):
    """{header name -> column index} for the first occurrence of each named header."""
    col = {}
    for idx, h in enumerate(header_row):
        h = clean(h)
        if h and h not in col:
            col[h] = idx
    return col


def parse_indicator_row(row, col):
    """One SoP indicator row -> a dict, regardless of which sheet shape it came from.

    'Benchmark specified for the standard' and 'Reported data' are each followed, on most
    sheets, by an unnamed column holding a plain-English gloss of that value (e.g. '% of cases
    resolved within specified time/standard') — but not on every sheet. Since that column
    usually has no header, detect it as: the very next index that isn't itself a *named*
    column. MPWZ,MADHYA PRADESH is the one sheet that actually names it "Benchmark meaning"
    instead of leaving it blank, so that named header is checked first."""
    named_indices = set(col.values())

    def gloss_after(idx):
        if idx is None:
            return None
        nxt = idx + 1
        if nxt in named_indices or nxt >= len(row):
            return None
        return row[nxt]

    benchmark_idx = col.get('Benchmark specified for the standard')
    reported_idx = col.get('Reported data')
    benchmark_meaning = row[col['Benchmark meaning']] if 'Benchmark meaning' in col else gloss_after(benchmark_idx)
    reported_meaning = gloss_after(reported_idx)
    reported_raw = row[reported_idx] if reported_idx is not None else None
    return {
        'type': clean(row[col['Indicator Type']]) if 'Indicator Type' in col else None,
        'indicator': clean(row[col['Indicator']]) if 'Indicator' in col else None,
        'meaning': clean(row[col['Indicator meaning']]) if 'Indicator meaning' in col else None,
        'standard_specified': clean(row[col['Standard Specified']]) if 'Standard Specified' in col else None,
        'benchmark': to_number(row[benchmark_idx]) if benchmark_idx is not None else None,
        'benchmark_meaning': clean(benchmark_meaning),
        'reported': to_number(reported_raw),
        'reported_meaning': clean(reported_meaning),
        'comparison_possible': to_bool_yn(row[col['Comparison Possible?']]) if 'Comparison Possible?' in col else None,
        'standard_met': to_bool_yn(row[col['Standard met?']]) if 'Standard met?' in col else None,
        'reason_not_comparable': clean(row[col['If comparison not possible, why?']]) if 'If comparison not possible, why?' in col else None,
    }


def parse_discom_sheet(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    full_name = clean(rows[0][0]) if rows else None
    code, _, state_part = sheet_name.partition(',')
    code = code.strip()
    state = STATE_MAP.get(state_part.strip().upper(), state_part.strip() or 'Unknown')

    years = {}
    cur_reg = []
    i = 1
    while i < n:
        r = rows[i]
        if r is None or all(c is None for c in r):
            i += 1
            continue
        c0 = clean(r[0])
        if c0 == 'Year':
            col = header_cols(r)
            j = i + 1
            by_year = {}
            while j < n:
                rj = rows[j]
                if rj is None or all(c is None for c in rj):
                    j += 1
                    continue
                c0j = clean(rj[0])
                if c0j == 'Year':
                    break
                if c0j and YEAR_RE.match(str(c0j)):
                    by_year.setdefault(c0j, []).append(rj)
                    j += 1
                else:
                    break
            for year, ents in by_year.items():
                entries = [parse_indicator_row(e, col) for e in ents]
                years[year] = {
                    'regulation': ' | '.join(cur_reg) if cur_reg else None,
                    'indicators': entries,
                }
            cur_reg = []
            i = j
        else:
            # a regulation-citation row can list more than one amendment across separate
            # columns (e.g. row 1 of MPCZ: principal regulation in col 0, a Revision-II amendment
            # in col 1, a further amendment in col 2) — collect every non-empty cell, not just
            # column 0, or later amendments silently go missing.
            cur_reg.extend(str(clean(c)) for c in r if clean(c) is not None)
            i += 1

    return {
        'sheet': sheet_name,
        'full_name': full_name,
        'short_name': code,
        'state': state,
        'years': years,
    }


def parse_framework_sheet(sheet_name, ws):
    """UP / TAMIL NADU / KARNATAKA / ANDHRA PRADESH / WEST BENGAL: a flat list of the regulation's
    named indicators with no Year column and no reported figures — just what's required, not
    whether it's met."""
    rows = list(ws.iter_rows(values_only=True))
    state = STATE_MAP.get(sheet_name.strip().upper(), sheet_name.strip())
    regulation = []
    header_idx = None
    for i, r in enumerate(rows):
        if r and clean(r[0]) == 'Indicator Type':
            header_idx = i
            break
        if r and any(c is not None for c in r):
            for c in r:
                if clean(c):
                    regulation.append(str(clean(c)))
    if header_idx is None:
        return {'state': state, 'regulation': ' | '.join(regulation), 'indicators': []}

    col = header_cols(rows[header_idx])
    entries = []
    for r in rows[header_idx + 1:]:
        if r is None or all(c is None for c in r):
            continue
        if not clean(r[col.get('Indicator Type', 0)]):
            continue
        entries.append(parse_indicator_row(r, col))

    return {'state': state, 'regulation': ' | '.join(regulation), 'indicators': entries}


discoms = []
frameworks = []
for sheet_name in wb.sheetnames:
    if sheet_name in EXCLUDE_SHEETS:
        continue
    if sheet_name in FRAMEWORK_ONLY_SHEETS:
        frameworks.append(parse_framework_sheet(sheet_name, wb[sheet_name]))
    else:
        discoms.append(parse_discom_sheet(sheet_name, wb[sheet_name]))

all_years = sorted({y for d in discoms for y in d['years']}, reverse=True)

out = {
    'state_order': STATE_ORDER,
    'years': all_years,
    'discoms': discoms,
    'frameworks': frameworks,
}

path = 'ui/public/data/state_specific.json'
with open(path, 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f'Wrote {path}')
print(f'{len(discoms)} DISCOMs across {len({d["state"] for d in discoms})} states, years={all_years}')
for d in discoms:
    y0 = all_years[0]
    y = d['years'].get(y0)
    n_indicators = len(y['indicators']) if y else 0
    print(f"  {d['short_name']:10} ({d['state']:16}) {y0}: {n_indicators} indicators")
print(f'{len(frameworks)} framework-only states: {[f["state"] for f in frameworks]}')
for fr in frameworks:
    print(f"  {fr['state']:16} indicators={len(fr['indicators'])}")
