import openpyxl, datetime, json, re

wb = openpyxl.load_workbook('data/PQ_Dashboard_ACPET.xlsx', data_only=True)

STATE_MAP = {
    'MSEDCL,MAHARASHTRA': 'Maharashtra',
    'DGCVL,GUJARAT': 'Gujarat',
    'MGVCL,GUJARAT': 'Gujarat',
    'UGVCL,GUJARAT': 'Gujarat',
    'PGVCL,GUJARAT': 'Gujarat',
    'MPEZ,MADHYA PRADESH': 'Madhya Pradesh',
    'MPWZ,MADHYA PRADESH': 'Madhya Pradesh',
    'MPCZ, MADHYA PRADESH': 'Madhya Pradesh',
    'AVVNL,RAJASTHAN': 'Rajasthan',
    'JVVNL,RAJASTHAN': 'Rajasthan',
    'JVVNL,RAJ': 'Rajasthan',
    'TPCODL,ODISHA': 'Odisha',
    'TPNODL,ODISHA': 'Odisha',
}

SHORT_MAP = {
    'MSEDCL,MAHARASHTRA': 'MSEDCL',
    'DGCVL,GUJARAT': 'DGVCL',
    'MGVCL,GUJARAT': 'MGVCL',
    'UGVCL,GUJARAT': 'UGVCL',
    'PGVCL,GUJARAT': 'PGVCL',
    'MPEZ,MADHYA PRADESH': 'MPEZ',
    'MPWZ,MADHYA PRADESH': 'MPWZ',
    'MPCZ, MADHYA PRADESH': 'MPCZ',
    'AVVNL,RAJASTHAN': 'AVVNL',
    'JVVNL,RAJASTHAN': 'JVVNL (Jaipur)',
    'JVVNL,RAJ': 'JdVVNL (Jodhpur)',
    'TPCODL,ODISHA': 'TPCODL',
    'TPNODL,ODISHA': 'TPNODL',
}

EXCLUDE_SHEETS = {'Sheet13', 'Sheet14'}

def to_hours(val, meaning, indicator):
    """Normalize a duration-type reading (SAIDI/CAIDI) to hours."""
    if val is None or val in ('N/A', 'N/A '):
        return None
    if isinstance(val, datetime.time):
        return round(val.hour + val.minute/60 + val.second/3600, 3)
    if isinstance(val, datetime.timedelta):
        return round(val.total_seconds()/3600, 3)
    if isinstance(val, (int, float)):
        m = (meaning or '').lower()
        if 'minute' in m:
            return round(val/60, 3)
        if 'hour' in m or 'hr' in m:
            return round(val, 3)
        # fallback: CAIDI/SAIDI raw numbers without explicit unit in meaning
        return round(val, 3)
    return None

def to_count(val):
    if val is None or isinstance(val, str):
        return None
    if isinstance(val, (int, float)):
        return round(val, 3)
    return None

def to_pct(val):
    """Normalize a 0-1 fraction or 0-100 number to a 0-100 percentage. Flags anomalies."""
    if val is None or isinstance(val, str):
        return None, False
    if isinstance(val, (int, float)):
        if val <= 1.0:
            return round(val*100, 3), False
        else:
            return round(val, 3), True  # flagged: already on 0-100 scale in a field where others use fraction
    return None, False

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return None
        return v
    if isinstance(v, datetime.time):
        return v.strftime('%H:%M')
    if isinstance(v, datetime.timedelta):
        total_h = v.total_seconds()/3600
        return f'{total_h:.2f} hrs (raw timedelta)'
    return v

discoms = []

for sheet_name in wb.sheetnames:
    if sheet_name in EXCLUDE_SHEETS:
        continue.
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    full_name = rows[0][0]
    regulation_lines = []
    header_idx = None
    for i, r in enumerate(rows[1:], start=1):
        if r[0] == 'Year':
            header_idx = i
            break
        if r[0]:
            regulation_lines.append(str(r[0]).strip())
    if header_idx is None:
        continue
    header = rows[header_idx]
    data_rows = rows[header_idx+1:]

    indicators = []
    for r in data_rows:
        if not r or r[2] is None:
            continue
        # stop if we hit a second header block (PGVCL sheet has a duplicated block)
        if r[0] == 'Year':
            break
        indicator_type = clean(r[1])
        indicator = clean(r[2])
        indicator_meaning = clean(r[3])
        standard_specified = clean(r[4])
        benchmark = clean(r[5])
        benchmark_meaning = clean(r[6])
        reported_raw_val = r[7]
        reported_meaning = clean(r[8])
        comparison_possible = clean(r[9])
        standard_met = clean(r[10])
        reason_not_comparable = clean(r[11]) if len(r) > 11 else None

        entry = {
            'type': indicator_type,
            'indicator': indicator,
            'indicator_meaning': indicator_meaning,
            'standard_specified': standard_specified,
            'benchmark': clean(benchmark) if not isinstance(benchmark, (int,float)) else benchmark,
            'benchmark_meaning': benchmark_meaning,
            'reported_raw': clean(reported_raw_val),
            'reported_meaning': reported_meaning,
            'comparison_possible': comparison_possible,
            'standard_met': standard_met,
            'reason_not_comparable': reason_not_comparable,
        }

        norm_indicator = (indicator or '').upper()
        flagged = False
        if norm_indicator in ('SAIDI',):
            entry['value_hours'] = to_hours(reported_raw_val, reported_meaning, indicator)
        elif norm_indicator in ('CAIDI',):
            entry['value_hours'] = to_hours(reported_raw_val, reported_meaning, indicator)
        elif norm_indicator in ('SAIFI', 'MAIFI'):
            entry['value_count'] = to_count(reported_raw_val)
        elif norm_indicator in ('VOLTAGE VARIATION', 'HARMONICS'):
            pct, flagged = to_pct(reported_raw_val)
            entry['value_pct'] = pct
            entry['unit_anomaly'] = flagged
        elif 'TRANSFORMER FAILURE' in norm_indicator or 'DISTRIBUTION TRANSFORMER' in norm_indicator:
            pct, flagged = to_pct(reported_raw_val)
            entry['value_pct'] = pct
            entry['unit_anomaly'] = flagged
        elif 'BILLING' in norm_indicator:
            pct, flagged = to_pct(reported_raw_val)
            entry['value_pct'] = pct
            entry['unit_anomaly'] = flagged

        indicators.append(entry)

    discoms.append({
        'sheet': sheet_name,
        'full_name': full_name,
        'short_name': SHORT_MAP.get(sheet_name, sheet_name.split(',')[0]),
        'state': STATE_MAP.get(sheet_name, 'Unknown'),
        'regulation': ' | '.join(regulation_lines),
        'indicators': indicators,
    })

# Comparability score per DISCOM: purely derived from the sheet's own columns.
# - Standards Availability: % of indicators with a real standard specified (not N/A)
# - Comparability: % of indicators where Comparison Possible == Yes
# - Compliance (of comparable ones): % where Standard met == Yes
for d in discoms:
    inds = d['indicators']
    n = len(inds)
    def is_na(x):
        return x is None or str(x).strip().upper() in ('N/A', 'NA', '')
    has_standard = sum(1 for i in inds if not is_na(i['standard_specified']))
    reported = sum(1 for i in inds if i['reported_raw'] is not None)

    # An indicator only truly counts as "comparable"/"compliant" if a value was
    # actually reported. Some sheets mark Comparison Possible=Yes / Standard Met=Yes
    # even though the reported-data cell is blank -- that's a reporting-integrity
    # gap, not real comparability, so it must not inflate the score.
    claimed_comparable = [i for i in inds if str(i['comparison_possible']).strip().upper() == 'YES']
    comparable_inds = [i for i in claimed_comparable if i['reported_raw'] is not None]
    comparable = len(comparable_inds)
    phantom_comparable = [i for i in claimed_comparable if i['reported_raw'] is None]
    met = sum(1 for i in comparable_inds if str(i['standard_met']).strip().upper() == 'YES')

    d['scoring'] = {
        'total_indicators': n,
        'standards_available_pct': round(100*has_standard/n, 1) if n else 0,
        'data_reported_pct': round(100*reported/n, 1) if n else 0,
        'comparable_pct': round(100*comparable/n, 1) if n else 0,
        'compliance_pct': round(100*met/len(comparable_inds), 1) if comparable_inds else None,
        'composite_score': round(
            (100*has_standard/n)*0.25 + (100*reported/n)*0.25 + (100*comparable/n)*0.30 +
            ((100*met/len(comparable_inds)) if comparable_inds else 0)*0.20, 1
        ) if n else 0,
        'phantom_comparable_count': len(phantom_comparable),
        'phantom_comparable_indicators': [i['indicator'] for i in phantom_comparable],
    }
    score = d['scoring']['composite_score']
    d['scoring']['grade'] = 'A' if score >= 70 else ('B' if score >= 45 else 'C')

with open('data/discoms.json', 'w', encoding='utf-8') as f:
    json.dump(discoms, f, indent=2, ensure_ascii=False)

print(f'Extracted {len(discoms)} DISCOMs')
for d in discoms:
    print(f"  {d['short_name']:20} ({d['state']:16}) - {len(d['indicators'])} indicators, score={d['scoring']['composite_score']}, grade={d['scoring']['grade']}")
