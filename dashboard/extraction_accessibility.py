"""Extract data/ACCESSIBILITY.xlsx into ui/public/data/accessibility.json for the Next.js app.

Unlike Common Indicators.xlsx (per-DISCOM sheets, 5 years of numeric indicators each), this
workbook is flat and year-less — two simple sheets:
  - REGULATION: one row per state, whether that state's SERC regulation is published online.
  - REPORTED DATAPERFORMANCE: one row per DISCOM, whether its reported performance data is
    published on the SERC website and whether that publication is machine-readable.
Written straight to ui/public/data/ (not data/, unlike discoms2.json) since there's no existing
build step that copies repo-root data/ output into ui/public/data/ — this way the app can fetch
it directly with nothing to remember to sync.
"""
import json
import openpyxl

wb = openpyxl.load_workbook('data/ACCESSIBILITY.xlsx', data_only=True)

# Same canonical state names/order as extraction_common.py's STATE_MAP/STATE_ORDER, so this
# dataset lines up with the rest of the app (map state names, discoms2.json's state_order).
STATE_MAP = {
    'MAHARASHTRA': 'Maharashtra', 'GUJARAT': 'Gujarat', 'RAJASTHAN': 'Rajasthan',
    'ODISHA': 'Odisha', 'TELANGANA': 'Telangana', 'MADHYA PRADESH': 'Madhya Pradesh',
    'KARNATAKA': 'Karnataka', 'TAMIL NADU': 'Tamil Nadu', 'BIHAR': 'Bihar',
    'WEST BENGAL': 'West Bengal', 'UTTAR PRADESH': 'Uttar Pradesh', 'ANDHRA PRADESH': 'Andhra Pradesh',
}
STATE_ORDER = ['Maharashtra', 'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Odisha', 'Telangana',
               'Karnataka', 'Tamil Nadu', 'Bihar', 'West Bengal', 'Uttar Pradesh', 'Andhra Pradesh']


def to_bool(v):
    """'Yes'/'No' -> True/False; anything else ('N/A', blank, None) -> None (not applicable)."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s == 'yes':
        return True
    if s == 'no':
        return False
    return None


def norm_state(s):
    return STATE_MAP.get(str(s).strip().upper(), str(s).strip())


# ---------- REGULATION: one row per state ----------
ws = wb['REGULATION']
rows = list(ws.iter_rows(values_only=True))[1:]  # skip header row
states = []
for state, available in rows:
    if state is None:
        continue
    states.append({
        'state': norm_state(state),
        'regulation_available': to_bool(available),
    })
states.sort(key=lambda s: STATE_ORDER.index(s['state']) if s['state'] in STATE_ORDER else 99)

# ---------- REPORTED DATAPERFORMANCE: one row per DISCOM ----------
ws = wb['REPORTED DATAPERFORMANCE']
rows = list(ws.iter_rows(values_only=True))[1:]  # skip header row
discoms = []
for row in rows:
    full_name, abbrev, state, available_on_serc, machine_readable = row[:5]
    drive_link = row[5] if len(row) > 5 else None
    if full_name is None:
        continue
    discoms.append({
        'discom': str(full_name).strip(),
        'abbreviation': str(abbrev).strip(),
        'state': norm_state(state),
        'available_on_serc': to_bool(available_on_serc),
        'machine_readable': to_bool(machine_readable),
        'drive_link': str(drive_link).strip() if drive_link else None,
    })
discoms.sort(key=lambda d: (STATE_ORDER.index(d['state']) if d['state'] in STATE_ORDER else 99, d['discom']))

summary = {
    'states_total': len(states),
    'states_regulation_available': sum(1 for s in states if s['regulation_available'] is True),
    'discoms_total': len(discoms),
    'discoms_available_on_serc': sum(1 for d in discoms if d['available_on_serc'] is True),
    'discoms_machine_readable': sum(1 for d in discoms if d['machine_readable'] is True),
}

out = {
    'state_order': STATE_ORDER,
    'states': states,
    'discoms': discoms,
    'summary': summary,
}

path = 'ui/public/data/accessibility.json'
with open(path, 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f'Wrote {path}')
print(json.dumps(summary, indent=2))
